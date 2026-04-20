import traceback

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .utils import br_display, br_now


def _cell_style(c, font=None, fill=None, align=None, border=None):
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border


_THIN_BORDER = Border(**{s: Side(style="thin", color="CCCCCC")
                         for s in ("left", "right", "bottom")})
_DATA_BORDER = Border(**{s: Side(style="thin", color="DDDDDD")
                         for s in ("left", "right", "bottom")})
_DATA_FONT = Font(name="Arial", size=10)
_DATA_FONT_DARK = Font(name="Arial", size=10, color="000000")
_WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
_CONF_FILL = {
    "alta": PatternFill("solid", start_color="E2EFDA"),
    "media": PatternFill("solid", start_color="FFF2CC"),
    "baixa": PatternFill("solid", start_color="FCE4D6"),
}


class ExcelReportManager:
    def __init__(self, report_path: str):
        self.report_path = report_path

    def init_excel(self):
        from os import path

        if path.exists(self.report_path):
            print(f"📊 Relatório existente: {self.report_path}")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Anúncios Detectados"
        headers = ["Data/Hora", "Rádio", "Anunciante", "Produto/Serviço",
                   "Confiança", "Trecho do Anúncio", "Arquivo de Áudio"]
        h_fill = PatternFill("solid", start_color="1F4E79")
        h_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            _cell_style(c, font=h_font, fill=h_fill, align=h_align, border=_THIN_BORDER)
        ws.row_dimensions[1].height = 60
        ws.freeze_panes = "A2"
        for i, w in enumerate([20, 15, 22, 22, 12, 60, 40], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws2 = wb.create_sheet("Resumo por Rádio")
        h2_font = Font(bold=True, color="FFFFFF", name="Arial")
        h2_fill = PatternFill("solid", start_color="2E75B6")
        resumo_headers = [
            "Rádio", "Total de Anúncios", "Última Detecção",
            "Início da Sessão", "Fim da Sessão", "Duração da Sessão",
        ]
        for col, h in enumerate(resumo_headers, 1):
            _cell_style(ws2.cell(row=1, column=col, value=h), font=h2_font, fill=h2_fill)
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 20
        ws2.column_dimensions["C"].width = 22
        ws2.column_dimensions["D"].width = 22
        ws2.column_dimensions["E"].width = 22
        ws2.column_dimensions["F"].width = 22

        ws3 = wb.create_sheet("Resumo por Anunciante")
        for col, h in enumerate(["Anunciante", "Total de Anúncios", "Última Detecção"], 1):
            _cell_style(ws3.cell(row=1, column=col, value=h), font=h2_font, fill=h2_fill)
        ws3.column_dimensions["A"].width = 25
        ws3.column_dimensions["B"].width = 20
        ws3.column_dimensions["C"].width = 20

        ws.auto_filter.ref = ws.dimensions
        ws2.auto_filter.ref = ws2.dimensions
        ws3.auto_filter.ref = ws3.dimensions

        wb.save(self.report_path)
        print(f"📊 Relatório criado: {self.report_path}")

    def append_to_excel(self, station, info, audio_file, start_time, session_excel_rows):
        try:
            from os import path

            wb = load_workbook(self.report_path)
            ws = wb["Anúncios Detectados"]
            row = ws.max_row + 1
            conf = info.get("confianca", "baixa")
            fill = _CONF_FILL.get(conf, PatternFill("solid", start_color="FFFFFF"))
            values = [
                br_display(), station,
                info.get("anunciante") or "—",
                info.get("produto") or "—",
                conf.upper(),
                info.get("trecho") or "—",
                path.basename(audio_file),
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                _cell_style(c, font=_DATA_FONT, fill=fill, border=_DATA_BORDER,
                            align=Alignment(vertical="center", wrap_text=(col == 6)))
            ws.row_dimensions[row].height = 18

            ws2 = wb["Resumo por Rádio"]
            inicio_str = start_time.strftime("%d/%m/%Y %H:%M:%S") if start_time else "—"

            row_idx = session_excel_rows.get(station)
            if row_idx:
                sr = list(ws2.iter_rows(min_row=row_idx, max_row=row_idx))[0]
                sr[1].value = (sr[1].value or 0) + 1
                sr[2].value = br_display()
                for cell in sr:
                    cell.fill = _WHITE_FILL
                    cell.font = _DATA_FONT_DARK
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                nr = ws2.max_row + 1
                session_excel_rows[station] = nr
                for col, val in enumerate([station, 1, br_display(), inicio_str, "—", "—"], 1):
                    c = ws2.cell(row=nr, column=col, value=val)
                    c.fill = _WHITE_FILL
                    c.font = _DATA_FONT_DARK
                    c.alignment = Alignment(horizontal="center", vertical="center")

            anunciante_nome = info.get("anunciante")
            if anunciante_nome and anunciante_nome not in ("—", "Desconhecido"):
                if "Resumo por Anunciante" not in wb.sheetnames:
                    wb.create_sheet("Resumo por Anunciante")
                ws3 = wb["Resumo por Anunciante"]

                if ws3.max_row == 1 and not ws3.cell(row=1, column=1).value:
                    h2_font = Font(bold=True, color="FFFFFF", name="Arial")
                    h2_fill = PatternFill("solid", start_color="2E75B6")
                    for col, h in enumerate(["Anunciante", "Total de Anúncios", "Última Detecção"], 1):
                        _cell_style(ws3.cell(row=1, column=col, value=h), font=h2_font, fill=h2_fill)
                    ws3.column_dimensions["A"].width = 25
                    ws3.column_dimensions["B"].width = 20
                    ws3.column_dimensions["C"].width = 20

                sr3 = next((r for r in ws3.iter_rows(min_row=2) if r[0].value == anunciante_nome), None)
                if sr3:
                    sr3[1].value = (sr3[1].value or 0) + 1
                    sr3[2].value = br_display()
                    for cell in sr3:
                        cell.fill = _WHITE_FILL
                        cell.font = _DATA_FONT_DARK
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    nr3 = ws3.max_row + 1
                    for col, val in enumerate([anunciante_nome, 1, br_display()], 1):
                        c = ws3.cell(row=nr3, column=col, value=val)
                        c.fill = _WHITE_FILL
                        c.font = _DATA_FONT_DARK
                        c.alignment = Alignment(horizontal="center", vertical="center")
                ws3.auto_filter.ref = ws3.dimensions

            ws.auto_filter.ref = ws.dimensions
            ws2.auto_filter.ref = ws2.dimensions

            wb.save(self.report_path)
            print(f"  ✅ Excel salvo ({row - 1} anúncios)")
        except Exception as e:
            print(f"  ⚠️  Erro ao salvar Excel: {e}")
            traceback.print_exc()

    def finalize_session_excel(self, start_time, session_excel_rows):
        if not start_time or not session_excel_rows:
            return
        try:
            wb = load_workbook(self.report_path)
            ws2 = wb["Resumo por Rádio"]
            fim = br_now()
            fim_str = fim.strftime("%d/%m/%Y %H:%M:%S")
            duracao = fim - start_time
            total_s = int(duracao.total_seconds())
            h, rem = divmod(total_s, 3600)
            m, s = divmod(rem, 60)
            dur_str = f"{h:02d}h {m:02d}m {s:02d}s"

            for _, row_idx in session_excel_rows.items():
                row = list(ws2.iter_rows(min_row=row_idx, max_row=row_idx))[0]
                row[4].value = fim_str
                row[5].value = dur_str
                for cell in row:
                    cell.fill = _WHITE_FILL
                    cell.font = _DATA_FONT_DARK
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save(self.report_path)
            print(
                f"  📋 Sessão finalizada: início={start_time.strftime('%H:%M:%S')} "
                f"fim={fim.strftime('%H:%M:%S')} duração={dur_str}"
            )
        except Exception as e:
            print(f"  ⚠️  Erro ao finalizar sessão no Excel: {e}")
            traceback.print_exc()
